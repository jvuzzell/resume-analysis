import os
import pandas as pd 
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

current_directory = os.getcwd()

def apply_formatting(sheet):
    # Apply gray background to the first row
    for cell in sheet["1"]:
        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Make each column fit the width of the content
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                if "\n" in str(cell.value):
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # Enable line breaks
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Adjust row height to fit the content
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

def csv_to_excel(csv_folder, output_excel):
    all_dataframes = {}
    
    # Extract all valid CSV files first
    csv_files = [file for file in os.listdir(csv_folder) if file.endswith('.csv')]

    # Sort files based on the pg-* prefix
    sorted_files = sorted(csv_files, key=lambda x: int(x.split('_')[0].replace('pg-', '')))
    
    for file in sorted_files:
        csv_file = os.path.join(csv_folder, file)
        df = pd.read_csv(csv_file)
        # Use the file name (without extension) as the sheet name
        sheet_name = os.path.splitext(file)[0]
        all_dataframes[sheet_name] = df

    if all_dataframes:
        # Create a new Excel workbook
        workbook = Workbook()

        for sheet_name, df in all_dataframes.items():
            if not df.empty:  # Check if the DataFrame is empty
                # Add a new sheet to the workbook
                sheet = workbook.create_sheet(title=sheet_name)

                # Write DataFrame to the sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    sheet.append(row)

                # Apply formatting to the sheet
                apply_formatting(sheet)

                # Enable auto-filter to make columns sortable
                sheet.auto_filter.ref = sheet.dimensions

        # Remove the default sheet (Sheet) from the workbook
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        # Save the Excel workbook to the specified file
        workbook.save(output_excel)

        print("CSV reports have been added to separate sheets in the Excel spreadsheet.")
    else:
        print("No CSV reports found in the directory.")

if __name__ == "__main__":
    data_directory = os.path.join(current_directory, "data/") # Replace with the actual path to the directory containing CSV files
    output_excel_path = "reports/resume_review.xlsx"  # Replace with the desired output Excel file path

    csv_to_excel(data_directory, output_excel_path)
